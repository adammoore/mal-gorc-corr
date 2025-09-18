"""
GORC-MaLDReTH Correlation Table Export Application

This Flask application provides functionality to export the GORC-MaLDReTH 
correlation mapping tables to Excel and Google Sheets formats.

Author: Research Data Lifecycle Team
Version: 1.1.0
"""

import os
import json
import io
from datetime import datetime
from flask import Flask, render_template, send_file, jsonify, request, redirect, url_for, session
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from typing import Dict, List, Tuple, Any

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Enable CORS for API endpoints
CORS(app, resources={r"/api/*": {"origins": "*"}})

# MaLDReTH lifecycle stages
MALDRETH_STAGES = [
    'CONCEPTUALIZE', 'PLAN', 'COLLECT', 'PROCESS', 'ANALYSE', 
    'STORE', 'PUBLISH', 'PRESERVE', 'SHARE', 'ACCESS', 'TRANSFORM'
]

# GORC service categories
GORC_CATEGORIES = [
    'Research Object Repositories',
    'Discovery Services',
    'Services and Tools for Direct Research Tasks',
    'Services and Tools that Enable Workflows and Middleware',
    'Vocabulary and Semantic Object Services',
    'Commons Catalogues of All Services and Tools',
    'Persistent Identifier Services',
    'Security and Identification Services (AAI)',
    'Helpdesk Services'
]

# Revised GORC categories based on working group discussion (Sept 2025)
REVISED_GORC_CATEGORIES = [
    'Discovery Services',
    'Data Repositories',
    'Publication Repositories',
    'Documentation Repositories',
    'Software Repositories',
    'Data Deposit Services',
    'Quality Assurance Services',
    'API Connection Services',
    'Services and Tools that Enable Workflows and Middleware',
    'Vocabulary and Semantic Object Services',
    'Commons Catalogues of All Services and Tools',
    'Persistent Identifier Services',
    'Security and Identification Services (AAI)',
    'Helpdesk Services'
]

# Correlation data structure with separated markers and descriptions
CORRELATION_DATA = {
    'Research Object Repositories': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': '', 'description': ''},
        'COLLECT': {'marker': 'XX', 'description': 'Electronic lab notebooks enable data capture; Data repositories store collected datasets'},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': 'X', 'description': 'Primary storage infrastructure with versioning'},
        'PUBLISH': {'marker': 'X', 'description': 'Publication platforms with DOI assignment'},
        'PRESERVE': {'marker': 'X', 'description': 'Long-term preservation infrastructure'},
        'SHARE': {'marker': 'X', 'description': 'Sharing through repository interfaces'},
        'ACCESS': {'marker': 'X', 'description': 'Access via download and API mechanisms'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Discovery Services': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Literature and dataset discovery for hypothesis formation'},
        'PLAN': {'marker': 'X', 'description': 'Discovery of existing studies and methodologies'},
        'COLLECT': {'marker': '', 'description': ''},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': '', 'description': ''},
        'PUBLISH': {'marker': '', 'description': ''},
        'PRESERVE': {'marker': '', 'description': ''},
        'SHARE': {'marker': 'X', 'description': 'Resource discovery and recommendation'},
        'ACCESS': {'marker': 'X', 'description': 'Search interfaces for resource access'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Services and Tools for Direct Research Tasks': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Concept mapping and visualization tools'},
        'PLAN': {'marker': 'X', 'description': 'Project planning and DMP tools'},
        'COLLECT': {'marker': 'X', 'description': 'Data collection instruments and ELNs'},
        'PROCESS': {'marker': 'X', 'description': 'Data processing and cleaning tools'},
        'ANALYSE': {'marker': 'X', 'description': 'Statistical and visualization software'},
        'STORE': {'marker': '', 'description': ''},
        'PUBLISH': {'marker': '', 'description': ''},
        'PRESERVE': {'marker': '', 'description': ''},
        'SHARE': {'marker': '', 'description': ''},
        'ACCESS': {'marker': '', 'description': ''},
        'TRANSFORM': {'marker': 'X', 'description': 'Data integration and transformation tools'}
    },
    'Services and Tools that Enable Workflows and Middleware': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': '', 'description': ''},
        'COLLECT': {'marker': 'X', 'description': 'Automated collection workflows'},
        'PROCESS': {'marker': 'X', 'description': 'Processing pipeline orchestration'},
        'ANALYSE': {'marker': 'X', 'description': 'Analysis workflow coordination'},
        'STORE': {'marker': 'X', 'description': 'Storage workflow automation'},
        'PUBLISH': {'marker': '', 'description': ''},
        'PRESERVE': {'marker': '', 'description': ''},
        'SHARE': {'marker': '', 'description': ''},
        'ACCESS': {'marker': '', 'description': ''},
        'TRANSFORM': {'marker': 'X', 'description': 'Transformation workflow orchestration'}
    },
    'Vocabulary and Semantic Object Services': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': 'X', 'description': 'Terminology standardization for planning'},
        'COLLECT': {'marker': 'X', 'description': 'Controlled vocabularies for annotation'},
        'PROCESS': {'marker': 'X', 'description': 'Semantic harmonization frameworks'},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': 'X', 'description': 'Semantic metadata for organization'},
        'PUBLISH': {'marker': 'X', 'description': 'Vocabulary standards for discoverability'},
        'PRESERVE': {'marker': 'X', 'description': 'Semantic metadata preservation'},
        'SHARE': {'marker': '', 'description': ''},
        'ACCESS': {'marker': 'X', 'description': 'Semantic search capabilities'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Commons Catalogues of All Services and Tools': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Tool and repository discovery'},
        'PLAN': {'marker': 'X', 'description': 'Service selection and planning'},
        'COLLECT': {'marker': '', 'description': ''},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': '', 'description': ''},
        'PUBLISH': {'marker': '', 'description': ''},
        'PRESERVE': {'marker': '', 'description': ''},
        'SHARE': {'marker': '', 'description': ''},
        'ACCESS': {'marker': 'X', 'description': 'Service discovery interfaces'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Persistent Identifier Services': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': '', 'description': ''},
        'COLLECT': {'marker': '', 'description': ''},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': 'X', 'description': 'Stable identifiers for storage'},
        'PUBLISH': {'marker': 'X', 'description': 'DOI assignment for citation'},
        'PRESERVE': {'marker': 'X', 'description': 'Long-term identifier persistence'},
        'SHARE': {'marker': 'X', 'description': 'Reliable referencing for sharing'},
        'ACCESS': {'marker': 'X', 'description': 'PID resolution for access'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Security and Identification Services (AAI)': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Secure access to planning environments'},
        'PLAN': {'marker': 'X', 'description': 'Authentication for planning platforms'},
        'COLLECT': {'marker': 'X', 'description': 'Secure instrument and data access'},
        'PROCESS': {'marker': 'X', 'description': 'Processing platform authentication'},
        'ANALYSE': {'marker': 'X', 'description': 'Computational resource security'},
        'STORE': {'marker': 'X', 'description': 'Storage access control'},
        'PUBLISH': {'marker': 'X', 'description': 'Publication identity management'},
        'PRESERVE': {'marker': 'X', 'description': 'Preservation system security'},
        'SHARE': {'marker': 'X', 'description': 'Sharing permission management'},
        'ACCESS': {'marker': 'X', 'description': 'Resource access authentication'},
        'TRANSFORM': {'marker': 'X', 'description': 'Integration platform security'}
    },
    'Helpdesk Services': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Conceptualization guidance and support'},
        'PLAN': {'marker': 'X', 'description': 'Planning methodology assistance'},
        'COLLECT': {'marker': 'X', 'description': 'Collection troubleshooting support'},
        'PROCESS': {'marker': 'X', 'description': 'Processing technical assistance'},
        'ANALYSE': {'marker': 'X', 'description': 'Analysis tool usage support'},
        'STORE': {'marker': 'X', 'description': 'Storage procedure guidance'},
        'PUBLISH': {'marker': 'X', 'description': 'Publication process assistance'},
        'PRESERVE': {'marker': 'X', 'description': 'Preservation planning support'},
        'SHARE': {'marker': 'X', 'description': 'Sharing configuration help'},
        'ACCESS': {'marker': 'X', 'description': 'Access problem resolution'},
        'TRANSFORM': {'marker': 'X', 'description': 'Integration challenge support'}
    }
}

# Revised correlation data based on working group discussion (Sept 2025)
# Decomposed Research Object Repositories into specific components
REVISED_CORRELATION_DATA = {
    'Discovery Services': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Literature and dataset discovery for hypothesis formation'},
        'PLAN': {'marker': 'X', 'description': 'Discovery of existing studies and methodologies'},
        'COLLECT': {'marker': '', 'description': ''},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': 'X', 'description': 'Finding comparative datasets and analytical resources'},
        'STORE': {'marker': '', 'description': ''},
        'PUBLISH': {'marker': '', 'description': ''},
        'PRESERVE': {'marker': '', 'description': ''},
        'SHARE': {'marker': 'X', 'description': 'Discovery interfaces for shared resources'},
        'ACCESS': {'marker': 'X', 'description': 'Search and discovery of accessible datasets'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Data Repositories': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': '', 'description': ''},
        'COLLECT': {'marker': 'XX', 'description': 'Primary data storage infrastructure'},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': 'XX', 'description': 'Core data storage with versioning'},
        'PUBLISH': {'marker': '', 'description': ''},
        'PRESERVE': {'marker': 'X', 'description': 'Long-term data preservation'},
        'SHARE': {'marker': 'X', 'description': 'Data sharing infrastructure'},
        'ACCESS': {'marker': 'X', 'description': 'Data access and download mechanisms'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Publication Repositories': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': '', 'description': ''},
        'COLLECT': {'marker': '', 'description': ''},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': '', 'description': ''},
        'PUBLISH': {'marker': 'XX', 'description': 'Academic publication platforms with DOI assignment'},
        'PRESERVE': {'marker': 'X', 'description': 'Long-term publication preservation'},
        'SHARE': {'marker': 'X', 'description': 'Publication sharing and dissemination'},
        'ACCESS': {'marker': 'X', 'description': 'Open access to published works'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Documentation Repositories': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Project documentation and protocols'},
        'PLAN': {'marker': 'X', 'description': 'Planning documents and methodological notes'},
        'COLLECT': {'marker': 'X', 'description': 'Data collection protocols and metadata'},
        'PROCESS': {'marker': 'X', 'description': 'Processing documentation and workflows'},
        'ANALYSE': {'marker': 'X', 'description': 'Analysis documentation and code'},
        'STORE': {'marker': 'X', 'description': 'Storage documentation and schemas'},
        'PUBLISH': {'marker': 'X', 'description': 'Publication support materials'},
        'PRESERVE': {'marker': 'X', 'description': 'Preservation documentation'},
        'SHARE': {'marker': 'X', 'description': 'Sharing protocols and guides'},
        'ACCESS': {'marker': 'X', 'description': 'Access documentation and user guides'},
        'TRANSFORM': {'marker': 'X', 'description': 'Transformation and integration guides'}
    },
    'Software Repositories': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': 'X', 'description': 'Research software planning and design tools'},
        'COLLECT': {'marker': 'X', 'description': 'Data collection software and scripts'},
        'PROCESS': {'marker': 'XX', 'description': 'Data processing and cleaning software'},
        'ANALYSE': {'marker': 'XX', 'description': 'Statistical and analytical software packages'},
        'STORE': {'marker': '', 'description': ''},
        'PUBLISH': {'marker': 'X', 'description': 'Publication generation software'},
        'PRESERVE': {'marker': 'X', 'description': 'Software preservation and containerization'},
        'SHARE': {'marker': 'X', 'description': 'Code sharing and collaboration platforms'},
        'ACCESS': {'marker': 'X', 'description': 'Software access and download'},
        'TRANSFORM': {'marker': 'XX', 'description': 'Data transformation and migration tools'}
    },
    'Data Deposit Services': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': '', 'description': ''},
        'COLLECT': {'marker': 'X', 'description': 'Initial data deposit mechanisms'},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': 'X', 'description': 'Secure data deposit workflows'},
        'PUBLISH': {'marker': '', 'description': ''},
        'PRESERVE': {'marker': 'X', 'description': 'Preservation-ready deposit services'},
        'SHARE': {'marker': 'X', 'description': 'Shared deposit and submission'},
        'ACCESS': {'marker': '', 'description': ''},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Quality Assurance Services': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': 'X', 'description': 'Quality planning and standards definition'},
        'COLLECT': {'marker': 'X', 'description': 'Data quality validation during collection'},
        'PROCESS': {'marker': 'XX', 'description': 'Data cleaning and quality assurance'},
        'ANALYSE': {'marker': 'X', 'description': 'Analytical quality controls'},
        'STORE': {'marker': 'X', 'description': 'Storage integrity and quality checks'},
        'PUBLISH': {'marker': 'X', 'description': 'Publication quality review'},
        'PRESERVE': {'marker': 'XX', 'description': 'Long-term integrity monitoring'},
        'SHARE': {'marker': 'X', 'description': 'Quality verification for sharing'},
        'ACCESS': {'marker': 'X', 'description': 'Access quality and reliability'},
        'TRANSFORM': {'marker': 'X', 'description': 'Transformation quality assurance'}
    },
    'API Connection Services': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': '', 'description': ''},
        'COLLECT': {'marker': 'X', 'description': 'API-based data collection'},
        'PROCESS': {'marker': 'X', 'description': 'Processing pipeline APIs'},
        'ANALYSE': {'marker': 'X', 'description': 'Analytical service APIs'},
        'STORE': {'marker': 'X', 'description': 'Storage API integration'},
        'PUBLISH': {'marker': 'X', 'description': 'Publication API services'},
        'PRESERVE': {'marker': 'X', 'description': 'Preservation API connections'},
        'SHARE': {'marker': 'X', 'description': 'Sharing API endpoints'},
        'ACCESS': {'marker': 'XX', 'description': 'Primary API access mechanisms'},
        'TRANSFORM': {'marker': 'X', 'description': 'Data transformation APIs'}
    },
    'Services and Tools that Enable Workflows and Middleware': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': 'X', 'description': 'Workflow planning and design tools'},
        'COLLECT': {'marker': 'X', 'description': 'Data collection orchestration'},
        'PROCESS': {'marker': 'XX', 'description': 'Automated processing workflows'},
        'ANALYSE': {'marker': 'XX', 'description': 'Analytical workflow management'},
        'STORE': {'marker': 'X', 'description': 'Storage workflow automation'},
        'PUBLISH': {'marker': 'X', 'description': 'Publication workflow tools'},
        'PRESERVE': {'marker': 'X', 'description': 'Preservation workflow services'},
        'SHARE': {'marker': 'X', 'description': 'Sharing workflow management'},
        'ACCESS': {'marker': 'X', 'description': 'Access workflow orchestration'},
        'TRANSFORM': {'marker': 'XX', 'description': 'Data transformation pipelines'}
    },
    'Vocabulary and Semantic Object Services': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Ontology and vocabulary discovery'},
        'PLAN': {'marker': 'X', 'description': 'Semantic planning and schema design'},
        'COLLECT': {'marker': 'X', 'description': 'Metadata and semantic annotation'},
        'PROCESS': {'marker': 'X', 'description': 'Semantic data processing'},
        'ANALYSE': {'marker': 'X', 'description': 'Semantic analysis and reasoning'},
        'STORE': {'marker': 'X', 'description': 'Semantic storage and indexing'},
        'PUBLISH': {'marker': 'X', 'description': 'Semantic publication metadata'},
        'PRESERVE': {'marker': 'X', 'description': 'Semantic preservation standards'},
        'SHARE': {'marker': 'X', 'description': 'Semantic interoperability'},
        'ACCESS': {'marker': 'X', 'description': 'Semantic search and discovery'},
        'TRANSFORM': {'marker': 'X', 'description': 'Semantic transformation services'}
    },
    'Commons Catalogues of All Services and Tools': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': '', 'description': ''},
        'COLLECT': {'marker': '', 'description': ''},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': '', 'description': ''},
        'PUBLISH': {'marker': '', 'description': ''},
        'PRESERVE': {'marker': '', 'description': ''},
        'SHARE': {'marker': '', 'description': ''},
        'ACCESS': {'marker': 'X', 'description': 'Service discovery and cataloguing'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Persistent Identifier Services': {
        'CONCEPTUALIZE': {'marker': '', 'description': ''},
        'PLAN': {'marker': '', 'description': ''},
        'COLLECT': {'marker': '', 'description': ''},
        'PROCESS': {'marker': '', 'description': ''},
        'ANALYSE': {'marker': '', 'description': ''},
        'STORE': {'marker': 'X', 'description': 'PID assignment for stored data'},
        'PUBLISH': {'marker': 'XX', 'description': 'DOI and identifier management'},
        'PRESERVE': {'marker': 'XX', 'description': 'Persistent citation and referencing'},
        'SHARE': {'marker': 'X', 'description': 'Shareable persistent identifiers'},
        'ACCESS': {'marker': 'X', 'description': 'PID-based access resolution'},
        'TRANSFORM': {'marker': '', 'description': ''}
    },
    'Security and Identification Services (AAI)': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Identity management for collaborative planning'},
        'PLAN': {'marker': 'X', 'description': 'Secure planning environment access'},
        'COLLECT': {'marker': 'X', 'description': 'Authenticated data collection'},
        'PROCESS': {'marker': 'X', 'description': 'Secure processing environments'},
        'ANALYSE': {'marker': 'X', 'description': 'Authenticated analytical access'},
        'STORE': {'marker': 'X', 'description': 'Secure storage authentication'},
        'PUBLISH': {'marker': 'X', 'description': 'Publisher authentication and authorization'},
        'PRESERVE': {'marker': 'X', 'description': 'Preservation access control'},
        'SHARE': {'marker': 'X', 'description': 'Controlled sharing permissions'},
        'ACCESS': {'marker': 'X', 'description': 'Authenticated data access'},
        'TRANSFORM': {'marker': 'X', 'description': 'Secure transformation services'}
    },
    'Helpdesk Services': {
        'CONCEPTUALIZE': {'marker': 'X', 'description': 'Research planning guidance'},
        'PLAN': {'marker': 'X', 'description': 'Methodology and planning support'},
        'COLLECT': {'marker': 'X', 'description': 'Data collection technical support'},
        'PROCESS': {'marker': 'X', 'description': 'Processing pipeline assistance'},
        'ANALYSE': {'marker': 'X', 'description': 'Analytical methodology support'},
        'STORE': {'marker': 'X', 'description': 'Storage and backup guidance'},
        'PUBLISH': {'marker': 'X', 'description': 'Publication process support'},
        'PRESERVE': {'marker': 'X', 'description': 'Digital preservation advice'},
        'SHARE': {'marker': 'X', 'description': 'Data sharing best practices'},
        'ACCESS': {'marker': 'X', 'description': 'Access troubleshooting and support'},
        'TRANSFORM': {'marker': 'X', 'description': 'Integration challenge support'}
    }
}

# Stage-specific service concentrations
STAGE_CONCENTRATIONS = {
    'CONCEPTUALIZE': {
        'primary': ['Discovery Services', 'Direct Research Tools'],
        'secondary': ['Commons Catalogues', 'AAI', 'Helpdesk']
    },
    'PLAN': {
        'primary': ['Direct Research Tools', 'Vocabulary Services'],
        'secondary': ['Discovery Services', 'Commons Catalogues', 'AAI', 'Helpdesk']
    },
    'COLLECT': {
        'primary': ['Direct Research Tools', 'Research Object Repositories'],
        'secondary': ['Workflow Services', 'Vocabulary Services', 'AAI', 'Helpdesk']
    },
    'PROCESS': {
        'primary': ['Direct Research Tools', 'Workflow Services'],
        'secondary': ['Vocabulary Services', 'AAI', 'Helpdesk']
    },
    'ANALYSE': {
        'primary': ['Direct Research Tools', 'Workflow Services'],
        'secondary': ['AAI', 'Helpdesk']
    },
    'STORE': {
        'primary': ['Research Object Repositories', 'PID Services'],
        'secondary': ['Workflow Services', 'Vocabulary Services', 'AAI', 'Helpdesk']
    },
    'PUBLISH': {
        'primary': ['Research Object Repositories', 'PID Services'],
        'secondary': ['Vocabulary Services', 'AAI', 'Helpdesk']
    },
    'PRESERVE': {
        'primary': ['Research Object Repositories', 'PID Services'],
        'secondary': ['Vocabulary Services', 'AAI', 'Helpdesk']
    },
    'SHARE': {
        'primary': ['Research Object Repositories', 'Discovery Services'],
        'secondary': ['PID Services', 'AAI', 'Helpdesk']
    },
    'ACCESS': {
        'primary': ['Research Object Repositories', 'Discovery Services'],
        'secondary': ['Commons Catalogues', 'PID Services', 'Vocabulary Services', 'AAI', 'Helpdesk']
    },
    'TRANSFORM': {
        'primary': ['Direct Research Tools', 'Workflow Services'],
        'secondary': ['AAI', 'Helpdesk']
    }
}


class ExcelExporter:
    """
    Handles the creation and formatting of Excel workbooks containing
    GORC-MaLDReTH correlation data.
    """
    
    def __init__(self):
        """Initialize the ExcelExporter with default styling configurations."""
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def create_correlation_matrix_sheet(self, wb: Workbook) -> None:
        """
        Create the main correlation matrix worksheet with comments.
        
        Args:
            wb: The openpyxl Workbook object to add the sheet to
        """
        ws = wb.active
        ws.title = "Correlation Matrix"
        
        # Write headers
        ws.cell(row=1, column=1, value="GORC Categories")
        for col, stage in enumerate(MALDRETH_STAGES, start=2):
            cell = ws.cell(row=1, column=col, value=stage)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write GORC categories and correlation data
        for row, category in enumerate(GORC_CATEGORIES, start=2):
            # Category name
            cell = ws.cell(row=row, column=1, value=category)
            cell.fill = self.subheader_fill
            cell.font = self.subheader_font
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Correlation data with comments
            for col, stage in enumerate(MALDRETH_STAGES, start=2):
                correlation = CORRELATION_DATA.get(category, {}).get(stage, {'marker': '', 'description': ''})
                marker = correlation.get('marker', '')
                description = correlation.get('description', '')
                
                cell = ws.cell(row=row, column=col, value=marker)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.cell_border
                
                # Add comment if there's a description
                if description:
                    comment = Comment(description, "GORC-MaLDReTH Tool")
                    comment.width = 300
                    comment.height = 100
                    cell.comment = comment
                
                # Color coding for correlations
                if marker == 'XX':
                    cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
                    cell.font = Font(bold=True, color="CC0000")
                elif marker == 'X':
                    cell.fill = PatternFill(start_color="E5F5E5", end_color="E5F5E5", fill_type="solid")
                    cell.font = Font(bold=True, color="008800")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        for col in range(2, len(MALDRETH_STAGES) + 2):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
            
    def create_analysis_summary_sheet(self, wb: Workbook) -> None:
        """
        Create the analysis summary worksheet.
        
        Args:
            wb: The openpyxl Workbook object to add the sheet to
        """
        ws = wb.create_sheet(title="Analysis Summary")
        
        # Cross-cutting services section
        row_num = 1
        cell = ws.cell(row=row_num, column=1, value="Cross-Cutting GORC Services")
        cell.font = Font(bold=True, size=14)
        row_num += 2
        
        cross_cutting_services = [
            ("AAI Services", "Universal security layer across ALL 11 MaLDReTH stages"),
            ("Helpdesk Services", "Comprehensive user support throughout ALL lifecycle activities"),
            ("Research Object Repositories", "Core infrastructure for 5 key stages (Store, Preserve, Publish, Share, Access)"),
            ("Direct Research Tools", "Primary enabler for active research phases (Conceptualize, Plan, Collect, Process, Analyse, Transform)")
        ]
        
        for service, description in cross_cutting_services:
            ws.cell(row=row_num, column=1, value=service).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=description)
            row_num += 1
        
        # Stage-specific concentrations
        row_num += 2
        cell = ws.cell(row=row_num, column=1, value="Stage-Specific Service Concentrations")
        cell.font = Font(bold=True, size=14)
        row_num += 2
        
        # Headers
        headers = ["MaLDReTH Stage", "Primary GORC Services", "Secondary GORC Services"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
        row_num += 1
        
        # Data
        for stage in MALDRETH_STAGES:
            ws.cell(row=row_num, column=1, value=stage)
            ws.cell(row=row_num, column=2, value=', '.join(STAGE_CONCENTRATIONS.get(stage, {}).get('primary', [])))
            ws.cell(row=row_num, column=3, value=', '.join(STAGE_CONCENTRATIONS.get(stage, {}).get('secondary', [])))
            row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        
    def create_key_findings_sheet(self, wb: Workbook) -> None:
        """
        Create the key findings worksheet.
        
        Args:
            wb: The openpyxl Workbook object to add the sheet to
        """
        ws = wb.create_sheet(title="Key Findings")
        
        row_num = 1
        cell = ws.cell(row=row_num, column=1, value="Key Findings")
        cell.font = Font(bold=True, size=14)
        row_num += 2
        
        findings = [
            ("Most Connected Services", "AAI (11 stages), Helpdesk (11 stages), Direct Research Tools (6 stages)"),
            ("Least Connected Services", "Commons Catalogues (3 stages), Discovery Services (4 stages)"),
            ("Repository-Centric Stages", "Store, Preserve, Publish, Share, Access all depend heavily on Research Object Repositories"),
            ("Tool-Intensive Stages", "Conceptualize through Analyse, plus Transform rely on Direct Research Tools"),
            ("Security & Support Universal", "Every stage requires authentication (AAI) and user assistance (Helpdesk)")
        ]
        
        for finding, description in findings:
            ws.cell(row=row_num, column=1, value=finding).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=description)
            row_num += 1
        
        # Add legend for correlation matrix
        row_num += 3
        ws.cell(row=row_num, column=1, value="Correlation Matrix Legend").font = Font(bold=True, size=12)
        row_num += 1
        ws.cell(row=row_num, column=1, value="X").font = Font(bold=True, color="008800")
        ws.cell(row=row_num, column=2, value="Standard correlation - Service supports this lifecycle stage")
        row_num += 1
        ws.cell(row=row_num, column=1, value="XX").font = Font(bold=True, color="CC0000")
        ws.cell(row=row_num, column=2, value="Strong correlation - Service is critical for this lifecycle stage")
        row_num += 1
        ws.cell(row=row_num, column=1, value="(empty)")
        ws.cell(row=row_num, column=2, value="No direct correlation between service and stage")
        
        # Add metadata
        row_num += 3
        ws.cell(row=row_num, column=1, value="Generated:").font = Font(italic=True)
        ws.cell(row=row_num, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80
        
    def generate_excel(self) -> io.BytesIO:
        """
        Generate the complete Excel workbook with all sheets.
        
        Returns:
            io.BytesIO: The Excel file as a bytes buffer
        """
        wb = Workbook()
        
        # Create all sheets
        self.create_correlation_matrix_sheet(wb)
        self.create_analysis_summary_sheet(wb)
        self.create_key_findings_sheet(wb)
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer


@app.route('/')
def index():
    """
    Render the main application page.
    
    Returns:
        str: Rendered HTML template
    """
    return render_template('index.html')

@app.route('/radial')
def radial_visualization():
    """
    Render the radial visualization page.
    
    Returns:
        str: Rendered HTML template for radial visualization
    """
    return render_template('radial_visualization.html')


@app.route('/api/radial-visualization-data')
def get_radial_visualization_data():
    """
    API endpoint to retrieve data formatted for radial visualization.
    
    Returns:
        Response: JSON response containing visualization data
    """
    try:
        # Prepare GORC categories with short names
        gorc_categories = [
            {'name': 'Research Object Repositories', 'shortName': 'Repositories'},
            {'name': 'Discovery Services', 'shortName': 'Discovery'},
            {'name': 'Services and Tools for Direct Research Tasks', 'shortName': 'Direct Tools'},
            {'name': 'Services and Tools that Enable Workflows and Middleware', 'shortName': 'Workflows'},
            {'name': 'Vocabulary and Semantic Object Services', 'shortName': 'Vocabulary'},
            {'name': 'Commons Catalogues of All Services and Tools', 'shortName': 'Catalogues'},
            {'name': 'Persistent Identifier Services', 'shortName': 'PIDs'},
            {'name': 'Security and Identification Services (AAI)', 'shortName': 'AAI'},
            {'name': 'Helpdesk Services', 'shortName': 'Helpdesk'}
        ]
        
        # Sample tool data for each stage
        stage_tools = {
            'CONCEPTUALIZE': [
                {'name': 'Miro', 'category': 'Mind Mapping'},
                {'name': 'Lucidchart', 'category': 'Diagramming'},
                {'name': 'XMind', 'category': 'Mind Mapping'}
            ],
            'PLAN': [
                {'name': 'DMP Tool', 'category': 'Data Management'},
                {'name': 'Trello', 'category': 'Project Planning'},
                {'name': 'Asana', 'category': 'Project Planning'}
            ],
            'COLLECT': [
                {'name': 'Open Data Kit', 'category': 'Data Collection'},
                {'name': 'SurveyMonkey', 'category': 'Surveys'},
                {'name': 'REDCap', 'category': 'Data Collection'}
            ],
            'PROCESS': [
                {'name': 'Jupyter', 'category': 'Computing'},
                {'name': 'RSpace', 'category': 'ELN'},
                {'name': 'LabArchives', 'category': 'ELN'}
            ],
            'ANALYSE': [
                {'name': 'SPSS', 'category': 'Statistics'},
                {'name': 'R Studio', 'category': 'Statistics'},
                {'name': 'Python', 'category': 'Computing'}
            ],
            'STORE': [
                {'name': 'Figshare', 'category': 'Repository'},
                {'name': 'Zenodo', 'category': 'Repository'},
                {'name': 'Dataverse', 'category': 'Repository'}
            ],
            'PUBLISH': [
                {'name': 'DRYAD', 'category': 'Repository'},
                {'name': 'GBIF', 'category': 'Domain Repository'},
                {'name': 'DataCite', 'category': 'Metadata'}
            ],
            'PRESERVE': [
                {'name': 'Archivematica', 'category': 'Archive'},
                {'name': 'Docker', 'category': 'Containers'},
                {'name': 'Preservica', 'category': 'Archive'}
            ],
            'SHARE': [
                {'name': 'GitHub', 'category': 'Code Sharing'},
                {'name': 'OSF', 'category': 'Collaboration'},
                {'name': 'Zenodo', 'category': 'Repository'}
            ],
            'ACCESS': [
                {'name': 'LDAP', 'category': 'Authentication'},
                {'name': 'Shibboleth', 'category': 'Authentication'},
                {'name': 'OAuth', 'category': 'Authentication'}
            ],
            'TRANSFORM': [
                {'name': 'Apache Spark', 'category': 'ETL'},
                {'name': 'Python', 'category': 'Programming'},
                {'name': 'Talend', 'category': 'ETL'}
            ]
        }
        
        response_data = {
            'stages': MALDRETH_STAGES,
            'gorcCategories': gorc_categories,
            'correlations': CORRELATION_DATA,
            'stageTools': stage_tools,
            'concentrations': STAGE_CONCENTRATIONS
        }
        
        return jsonify(response_data), 200
        
    except Exception as e:
        logger.error(f"Error generating radial visualization data: {str(e)}")
        return jsonify({'error': 'Failed to generate visualization data'}), 500


@app.route('/api/correlation-data')
def get_correlation_data():
    """
    API endpoint to retrieve correlation data in JSON format.
    
    Returns:
        Response: JSON response containing correlation data
    """
    try:
        response_data = {
            'stages': MALDRETH_STAGES,
            'categories': GORC_CATEGORIES,
            'correlations': CORRELATION_DATA,
            'concentrations': STAGE_CONCENTRATIONS
        }
        return jsonify(response_data), 200
    except Exception as e:
        logger.error(f"Error retrieving correlation data: {str(e)}")
        return jsonify({'error': 'Failed to retrieve correlation data'}), 500


@app.route('/export/excel')
def export_excel():
    """
    Export correlation data as an Excel file.
    
    Returns:
        Response: Excel file download
    """
    try:
        exporter = ExcelExporter()
        excel_buffer = exporter.generate_excel()
        
        filename = f"GORC_MaLDReTH_Correlation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
            cache_timeout=0
        )
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        return jsonify({'error': 'Failed to generate Excel file'}), 500


@app.route('/export/csv/<sheet_type>')
def export_csv(sheet_type: str):
    """
    Export specific sheet data as CSV.
    
    Args:
        sheet_type: Type of sheet to export ('matrix', 'summary', 'findings')
        
    Returns:
        Response: CSV file download
    """
    try:
        if sheet_type == 'matrix':
            # Create correlation matrix DataFrame with just markers
            df_data = []
            for category in GORC_CATEGORIES:
                row_data = {'GORC_Category': category}
                for stage in MALDRETH_STAGES:
                    correlation = CORRELATION_DATA.get(category, {}).get(stage, {'marker': ''})
                    row_data[stage] = correlation.get('marker', '')
                df_data.append(row_data)
            
            df = pd.DataFrame(df_data)
            filename = f"GORC_MaLDReTH_Matrix_{datetime.now().strftime('%Y%m%d')}.csv"
            
        elif sheet_type == 'summary':
            # Create summary DataFrame
            df_data = []
            for stage in MALDRETH_STAGES:
                row_data = {
                    'MaLDReTH_Stage': stage,
                    'Primary_Services': ', '.join(STAGE_CONCENTRATIONS.get(stage, {}).get('primary', [])),
                    'Secondary_Services': ', '.join(STAGE_CONCENTRATIONS.get(stage, {}).get('secondary', []))
                }
                df_data.append(row_data)
            
            df = pd.DataFrame(df_data)
            filename = f"GORC_MaLDReTH_Summary_{datetime.now().strftime('%Y%m%d')}.csv"
            
        else:
            return jsonify({'error': 'Invalid sheet type'}), 400
        
        # Convert DataFrame to CSV
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        csv_buffer.seek(0)
        
        # Convert to bytes
        csv_bytes = io.BytesIO()
        csv_bytes.write(csv_buffer.getvalue().encode('utf-8'))
        csv_bytes.seek(0)
        
        return send_file(
            csv_bytes,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting to CSV: {str(e)}")
        return jsonify({'error': 'Failed to generate CSV file'}), 500


# NEW ROUTES FOR REVISED STRUCTURE (Sept 2025)

@app.route('/revised')
def revised_matrix():
    """
    Render the revised correlation matrix page based on working group discussion.

    Returns:
        str: Rendered HTML template for revised matrix view
    """
    return render_template('revised_matrix.html')

@app.route('/revised-radial')
def revised_radial_visualization():
    """
    Render the revised radial visualization page.

    Returns:
        str: Rendered HTML template for revised radial visualization
    """
    return render_template('revised_radial_visualization.html')

@app.route('/api/revised-correlation-data')
def get_revised_correlation_data():
    """
    API endpoint to retrieve revised correlation data based on working group discussion.

    Returns:
        Response: JSON response containing revised correlation data
    """
    try:
        response_data = {
            'stages': MALDRETH_STAGES,
            'categories': REVISED_GORC_CATEGORIES,
            'correlations': REVISED_CORRELATION_DATA
        }
        return jsonify(response_data), 200
    except Exception as e:
        logger.error(f"Error retrieving revised correlation data: {str(e)}")
        return jsonify({'error': 'Failed to retrieve revised correlation data'}), 500

@app.route('/api/revised-radial-visualization-data')
def get_revised_radial_visualization_data():
    """
    API endpoint to retrieve data formatted for revised radial visualization.

    Returns:
        Response: JSON response containing revised visualization data
    """
    try:
        # Prepare revised GORC categories with short names
        revised_gorc_categories = [
            {'name': 'Discovery Services', 'shortName': 'Discovery'},
            {'name': 'Data Repositories', 'shortName': 'Data Repos'},
            {'name': 'Publication Repositories', 'shortName': 'Pub Repos'},
            {'name': 'Documentation Repositories', 'shortName': 'Doc Repos'},
            {'name': 'Software Repositories', 'shortName': 'Software'},
            {'name': 'Data Deposit Services', 'shortName': 'Deposit'},
            {'name': 'Quality Assurance Services', 'shortName': 'QA'},
            {'name': 'API Connection Services', 'shortName': 'APIs'},
            {'name': 'Services and Tools that Enable Workflows and Middleware', 'shortName': 'Workflows'},
            {'name': 'Vocabulary and Semantic Object Services', 'shortName': 'Vocabulary'},
            {'name': 'Commons Catalogues of All Services and Tools', 'shortName': 'Catalogues'},
            {'name': 'Persistent Identifier Services', 'shortName': 'PIDs'},
            {'name': 'Security and Identification Services (AAI)', 'shortName': 'AAI'},
            {'name': 'Helpdesk Services', 'shortName': 'Helpdesk'}
        ]

        response_data = {
            'stages': MALDRETH_STAGES,
            'categories': revised_gorc_categories,
            'correlations': REVISED_CORRELATION_DATA,
            'stageTools': {}  # Simplified for now
        }
        return jsonify(response_data), 200

    except Exception as e:
        logger.error(f"Error retrieving revised radial visualization data: {str(e)}")
        return jsonify({'error': 'Failed to retrieve revised radial visualization data'}), 500

@app.route('/comparison')
def comparison_view():
    """
    Render a comparison page showing both original and revised structures.

    Returns:
        str: Rendered HTML template for comparison view
    """
    return render_template('comparison.html')

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors."""
    return jsonify({'error': 'Resource not found'}), 404


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors."""
    logger.error(f"Internal server error: {str(error)}")
    return jsonify({'error': 'Internal server error'}), 500


if __name__ == '__main__':
    # Get port from environment variable for deployment
    port = int(os.environ.get('PORT', 5000))
    
    # Run the application
    app.run(
        host='0.0.0.0',
        port=port,
        debug=os.environ.get('FLASK_ENV') == 'development'
    )
